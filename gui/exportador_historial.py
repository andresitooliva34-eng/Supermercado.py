from datetime import datetime
from tkinter import filedialog, messagebox

# Librerías utilizadas para generar archivos PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)
from reportlab.lib.enums import TA_CENTER

# Librería utilizada para generar archivos Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# Colores utilizados en los archivos exportados
COLOR_TITULO = colors.HexColor("#2E7D32")
COLOR_ENCABEZADO = colors.HexColor("#2E7D32")
COLOR_FILA_ALTERNA = colors.HexColor("#F1F8E9")
COLOR_HEX_ENCABEZADO_EXCEL = "2E7D32"


def exportar_historial_pdf(ventana_padre, ventas):
    # Verifica que existan ventas para exportar
    if not ventas:
        messagebox.showwarning(
            "Atención",
            "No hay compras registradas para exportar."
        )
        return

    # Sugiere un nombre para el archivo usando la fecha actual
    nombre_sugerido = (
        f"historial_compras_"
        f"{datetime.now().strftime('%Y%m%d')}.pdf"
    )

    # Permite al usuario elegir dónde guardar el archivo
    ruta = filedialog.asksaveasfilename(
        parent=ventana_padre,
        title="Guardar historial como PDF",
        defaultextension=".pdf",
        filetypes=[("Archivo PDF", "*.pdf")],
        initialfile=nombre_sugerido
    )

    if not ruta:
        return

    try:
        # Genera el archivo PDF
        _generar_pdf(ventas, ruta)

        messagebox.showinfo(
            "Listo",
            f"Historial exportado correctamente a:\n{ruta}"
        )

    except Exception as e:
        # Muestra el error si ocurre un problema durante la exportación
        messagebox.showerror(
            "Error",
            f"No se pudo generar el PDF:\n{e}"
        )


def _generar_pdf(ventas, ruta_destino):
    # Configuración del documento PDF
    doc = SimpleDocTemplate(
        ruta_destino,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm
    )

    estilos = getSampleStyleSheet()

    # Estilo para el título
    estilo_titulo = ParagraphStyle(
        "Titulo",
        parent=estilos["Title"],
        textColor=COLOR_TITULO,
        fontSize=20,
        alignment=TA_CENTER
    )

    # Estilo para el subtítulo
    estilo_subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=estilos["Normal"],
        textColor=colors.HexColor("#558B2F"),
        fontSize=11,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # Información que aparecerá antes de la tabla
    elementos = [
        Paragraph(
            "Supermercado",
            estilo_titulo
        ),

        Paragraph(
            f"Historial de Compras &nbsp;|&nbsp; "
            f"{datetime.now().strftime('%d/%m/%Y - %H:%M')}",
            estilo_subtitulo
        ),
    ]

    # Columnas de la tabla
    encabezados = [
        "Venta",
        "Cliente",
        "Fecha",
        "Pago",
        "Total"
    ]

    # Convierte cada venta en una fila de la tabla
    filas = [
        [
            f"#{v['id']:03d}",
            v.get("cliente_nombre", ""),
            str(v.get("fecha", "")),
            v.get("metodo_pago", ""),
            f"${float(v['total']):,.2f}",
        ]
        for v in reversed(ventas)
    ]

    # Crea la tabla con los encabezados y las ventas
    tabla = Table(
        [encabezados] + filas,
        colWidths=[
            2.2 * cm,
            5 * cm,
            3.5 * cm,
            3 * cm,
            3 * cm
        ]
    )

    # Da formato visual a la tabla
    tabla.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_ENCABEZADO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            # Alterna los colores de las filas
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, COLOR_FILA_ALTERNA]),
        ])
    )

    elementos.append(tabla)

    # Calcula el total de todas las ventas
    total_general = sum(
        float(v["total"])
        for v in ventas
    )

    elementos.append(
        Spacer(1, 15)
    )

    estilo_total = ParagraphStyle(
        "Total",
        parent=estilos["Heading2"],
        textColor=COLOR_TITULO,
        fontSize=14,
        alignment=2
    )

    elementos.append(
        Paragraph(
            f"TOTAL GENERAL: ${total_general:,.2f}",
            estilo_total
        )
    )

    # Construye y guarda definitivamente el PDF
    doc.build(elementos)


def exportar_historial_excel(ventana_padre, ventas):
    # Verifica que existan ventas para exportar
    if not ventas:
        messagebox.showwarning(
            "Atención",
            "No hay compras registradas para exportar."
        )
        return

    # Sugiere un nombre para el archivo
    nombre_sugerido = (
        f"historial_compras_"
        f"{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    # Permite seleccionar dónde guardar el Excel
    ruta = filedialog.asksaveasfilename(
        parent=ventana_padre,
        title="Guardar historial como Excel",
        defaultextension=".xlsx",
        filetypes=[("Archivo Excel", "*.xlsx")],
        initialfile=nombre_sugerido
    )

    if not ruta:
        return

    try:
        # Genera el archivo Excel
        _generar_excel(ventas, ruta)

        messagebox.showinfo(
            "Listo",
            f"Historial exportado correctamente a:\n{ruta}"
        )

    except Exception as e:
        messagebox.showerror(
            "Error",
            f"No se pudo generar el Excel:\n{e}"
        )


def _generar_excel(ventas, ruta_destino):
    # Crea un nuevo archivo Excel
    libro = openpyxl.Workbook()

    # Selecciona la hoja principal
    hoja = libro.active
    hoja.title = "Historial de Compras"

    # Crea los encabezados
    encabezados = [
        "Venta",
        "Cliente",
        "Fecha",
        "Medio de Pago",
        "Total"
    ]

    hoja.append(encabezados)

    # Estilo del encabezado
    color_encabezado = PatternFill(
        start_color=COLOR_HEX_ENCABEZADO_EXCEL,
        end_color=COLOR_HEX_ENCABEZADO_EXCEL,
        fill_type="solid"
    )

    fuente_encabezado = Font(
        color="FFFFFF",
        bold=True
    )

    # Aplica el estilo a las celdas del encabezado
    for celda in hoja[1]:
        celda.fill = color_encabezado
        celda.font = fuente_encabezado
        celda.alignment = Alignment(
            horizontal="center"
        )

    # Agrega las ventas a la hoja
    for v in reversed(ventas):
        hoja.append([
            f"#{v['id']:03d}",
            v.get("cliente_nombre", ""),
            str(v.get("fecha", "")),
            v.get("metodo_pago", ""),
            float(v["total"]),
        ])

    # Calcula el total general
    total_general = sum(
        float(v["total"])
        for v in ventas
    )

    # Agrega el total al final de la tabla
    fila_total = hoja.max_row + 1

    hoja.cell(
        row=fila_total,
        column=4,
        value="TOTAL GENERAL:"
    ).font = Font(bold=True)

    hoja.cell(
        row=fila_total,
        column=5,
        value=total_general
    ).font = Font(bold=True)

    # Define el ancho de cada columna
    anchos = {
        "A": 12,
        "B": 25,
        "C": 18,
        "D": 16,
        "E": 14
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[columna].width = ancho

    # Formatea la columna de precios como moneda
    for fila in hoja.iter_rows(
        min_row=2,
        max_row=hoja.max_row,
        min_col=5,
        max_col=5
    ):
        for celda in fila:
            celda.number_format = '"$"#,##0.00'

    # Guarda el archivo Excel
    libro.save(ruta_destino)